from collections import defaultdict
from glob import glob

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from pathlib import Path
import logging
from datetime import datetime

from controllers.Payment import getDataFromPaymentsFile
from controllers.Receipt import receiptsToPDFs
from models.Member import Member
from utils import styles
from utils.loadSave import save


## Fonctions


def getNbMemberInList(email, name, surname, list):
    for nbMember, member in enumerate(list):
        if member.isThisMember(email, name, surname):
            return nbMember
    return False


def createOrResetMembersList(year):
    membersList = paths["listesAdherents"] / f"{year}.xlsx"
    if not Path(membersList).is_file():
        try:
            year = str(year)
            precYear = str(int(year)-1)
            nextYear = str(int(year)+1)

            workbook = Workbook()
            sheet = workbook.active

            fields = ["Adresse mail", "Nom", "Prénom", "IDs reçus", "Statut",
                      "Dern. adh.", "Régulier", "Envoi mails", "Adresse", "Code postal", "Ville", "Téléphone",
                      "Adh. "+precYear, "Adh. "+year, "Adh. "+nextYear,
                      "Dons "+year, "Total "+year, "Dern. paiement", "Dern. moy. paiement", "Remarques"]

            # Ajouter les en-têtes de colonnes
            sheet.append(fields)

            # repositionner le curseur
            sheet.cell(row=2, column=1)

            for nbCol, width in enumerate(styles.widthCols):
                col = sheet.column_dimensions[get_column_letter(nbCol + 1)]
                col.width = width

            # Appliquer le style de l'en-tête à la première ligne
            for cell in sheet[1]:
                cell.font = styles.fontHeader
                cell.border = styles.borderHeader
                cell.alignment = styles.Alignment(horizontal="center", vertical="center")

            # Sauvegarder le fichier
            workbook.save(membersList)
            workbook.close()
            logging.info(f"Fichier {membersList} créé")
            return sheet
        except OSError as error:
            logging.error(f"Impossible de créer le fichier {membersList} : {error}")
            return False
    else:  # Sinon remettre tout à zero
        try:
            workbook = load_workbook(membersList)
            sheet = workbook.active

            sheet.delete_rows(2, sheet.max_row)
            workbook.save(membersList)
            workbook.close()
            return sheet
        except OSError as error:
            logging.error(f"Impossible d'ouvrir le fichier {membersList} : {error}")
            return False


# Configurer Logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

## Créer les dossiers si besoin ##
currentPath = Path.cwd()
dataPath = currentPath / "donnees"

currentDatetime = datetime.now()

paths = {  # TODO : use Path()
    "actuel": dataPath,
    "listesAdherents": dataPath / "listesAdherents",
    "paiementsHelloAsso": dataPath / "paiements/helloAsso",
    "paiementsPaypal": dataPath / "paiements/paypal",
    "paiementsVirEspChq": dataPath / "paiements/virEspChq",
    "paiementsCB": dataPath / "paiements/CB",
    "recusFiscaux": dataPath / "recusFiscaux"
}

for key, path in paths.items():
    try:
        path.mkdir(parents=True, exist_ok=True)
        logging.info(f"Dossier créé ou déjà existant : {path}")
    except OSError as error:
        logging.error(f"Impossible de créer le dossier {path} : {error}")

paymentsFiles = {
    "helloAsso": glob(str(paths["paiementsHelloAsso"] / "*.xlsx")),
    "paypal": glob(str(paths["paiementsPaypal"] / "*.xlsx")),
    "virEspChq": glob(str(paths["paiementsVirEspChq"] / "*.xlsx")),
    "cb": glob(str(paths["paiementsCB"] / "*.csv"))
}

membersByYear = defaultdict(list)
years = []
for source, filePaths in paymentsFiles.items():
    for filePath in filePaths:
        for payment in getDataFromPaymentsFile(filePath, source):
            year = payment.year
            if year not in years:  # Si c'est la première fois à l'execution qu'on tombe sur cette année
                years.append(year)  # On ajoute cette année à la liste des années
                membersByYear[year] = []  # On créait une nouvelle liste d'adhérents pour cette année
                membersNotes = {}
                membersList = paths["listesAdherents"] / f"{year}.xlsx"
                if Path(membersList).is_file():
                    workbook = load_workbook(membersList)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        if row[19].value is not None:
                            membersNotes[row[0].value] = str(row[19].value)

                sheet = createOrResetMembersList(year)  # On créait ou on réinitialise le fichier liste des adhérents correspondant

                # On doit par conséquent créer le nouvel adhérent pour cette année
                newMember = Member(payment.email, payment.name, payment.surname, payment.address,
                                   payment.postalCode, payment.city, payment.phone) #args keys
                if payment.email in membersNotes:
                    newMember.notes = membersNotes[payment.email]
                newMember.addPayment(payment)
                membersByYear[year].append(newMember)
            else:  # S'il y a déjà une liste d'adhérents pour cette année
                nbMember = getNbMemberInList(payment.email, payment.name, payment.surname, membersByYear[year])
                if nbMember:  # S'il y a déjà cet adhérent dans la liste
                    # On le met à jour (si besoin)
                    member = membersByYear[year][nbMember]
                    member.updateContactData(payment.address, payment.postalCode, payment.city, payment.phone)
                    member.addPayment(payment)
                else:
                    # Sinon on le créait
                    newMember = Member(payment.email, payment.name, payment.surname, payment.address,
                                       payment.postalCode, payment.city, payment.phone)

                    if payment.email in membersNotes:
                        newMember.notes = membersNotes[payment.email]
                    newMember.addPayment(payment)
                    membersByYear[year].append(newMember)


#years.sort(reverse=True)
years.sort()

# nbReceipts = 0


for year in years:
    for member in membersByYear[year]:
        prevYearStr = str(int(year)-1)
        if membersByYear[prevYearStr] != []:
            nbMember = getNbMemberInList(member.email, member.name, member.surname, membersByYear[prevYearStr])
            if type(nbMember) is int:
                paidMembershipNextYear = membersByYear[prevYearStr][nbMember].amounts["paidMembershipNextYear"]
                if paidMembershipNextYear > 0:
                    member.amounts["paidMembershipLastYear"] = paidMembershipNextYear
                    member.status = "RAA"
                    member.lastMembership = int(year)-1

        total = member.amounts["totalYear"]
        restToPay = member.rate-member.amounts["paidMembershipLastYear"]
        membershipThisYear = min(restToPay, total)

        membershipNextYear = 0
        if datetime.strptime(member.receipts[-1].date, "%d/%m/%Y").month >= 9 and total >= member.rate:  # Après septembre et il reste des sous
            membershipNextYear = min(restToPay, total-membershipThisYear)

        totalDonation = total - membershipThisYear - membershipNextYear

        member.amounts["paidMembershipYear"] = membershipThisYear
        member.amounts["paidMembershipNextYear"] = membershipNextYear
        member.amounts["donationsYear"] = totalDonation

        if member.status == "NA":  # Vérifier si RA
            for yearPrec in years:
                nbMember = getNbMemberInList(member.email, member.name, member.surname, membersByYear[str(yearPrec)])
                if type(nbMember) is int:
                    if yearPrec < year:
                        member.lastMembership = yearPrec
                        member.status = "RA"


for year in years:
    membersByYear[year] = sorted(membersByYear[year], key=lambda member: member.lastPayment or datetime.min)
    workbookPath = paths["listesAdherents"] / f"{year}.xlsx"
    workbook = load_workbook(filename=workbookPath)
    sheet = workbook.active

    for member in membersByYear[year]:
        # Vérification que tous les adhérents ont des coordonnées de contact valides
        if member.hasValidAddress():
            # Création des reçus
            receiptsToPDFs(member.receipts)
            save.addMemberReceipts(member)

            # Sauvegarde des éventuelles remarques concernent un adhérent
            if member.notes is not None:
                save.addMemberNotes(member.email, member.notes)
        else:
            logging.warning(
                member.surname + ' ' + member.name + " n'a pas de coordonnées de contact valides. L'édition de ses reçus est impossible.")




        sheet.append(member.toArray())

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = styles.borderBody

    workbook.save(workbookPath)
    workbook.close()

save.save()
