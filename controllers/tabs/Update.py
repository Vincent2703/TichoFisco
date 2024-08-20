from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from models.Member import Member
from models.Save import Save
from utils.FileManager import initMembersFile, getDataFromPaymentsFile, exportMembersFile, exportMemberReceipts
from utils.LogManager import LogManager
from utils.PathManager import PathManager
from utils.misc import openDir


class Update:
    def __init__(self, view=None):
        self.view = view

        self.paths = PathManager().getPaths()
        self.paymentFiles = self.paths["paymentFilesPatterns"].items()

        self.currentStep = 1
        self.progressPercent = 0

        nbSteps = 0
        for files in self.paymentFiles:
            nbSteps += len(files[1])
        self.nbSteps = nbSteps

    def setView(self, view):
        self.view = view

    def openDataDir(self):
        dirPath = self.paths["actuel"]
        if not openDir(dirPath):
            LogManager().addLog("OS", LogManager.LOGTYPE_ERROR, f"Impossible d'ouvrir le dossier : '{dirPath}'")
            return False
        return True

    def processPayments(self):  # TODO: rename it
        def getNbMemberInList(email, name, surname, list):  # Permet de retrouver un adhérent parmi une liste d'instances
            for nbMember, member in enumerate(list):
                if member.isThisMember(email, name, surname):
                    return nbMember
            return False

        membersByYear = defaultdict(list)  # Contiendra la liste des membres pour chaque année
        years = []  # Liste des années

        """
        Dans un premier temps, on récupère les données des fichiers de paiements
        puis à partir de ces données, on enregistre les adhérents ainsi que leurs paiements
        """
        defaultRate = Save().defaultRate["value"]
        for source, filePaths in self.paymentFiles:  # Pour chaque fichier de paiement récupéré
            for filePath in filePaths:
                for payment in getDataFromPaymentsFile(filePath, source):  # On obtient ses données
                    year = payment.year
                    memberSavedData = {}
                    if year not in years:  # Si on n'est pas encore tombé sur l'année, on doit soit créer le fichier liste des adhérents correspondant, ou alors le réinit
                        years.append(year)
                        membersByYear[year] = []
                        membersList = self.paths["listesAdherents"] / f"{year}.xlsx"
                        if Path(membersList).is_file():  # Si le fichier existe déjà
                            workbook = load_workbook(membersList)
                            sheet = workbook.active
                            for row in sheet.iter_rows(min_row=2, max_row=len(sheet['A']) - 4):  # On regarde s'il y a des remarques à récupérer pour les remettre dans le fichier final si l'adh y est toujours présent
                                memberEmail = row[0].value
                                notes = row[20].value
                                rate = row[19].value
                                if notes is not None or rate != defaultRate:
                                    memberSavedData[memberEmail] = {"notes": '', "rate": defaultRate}
                                if notes is not None:
                                    memberSavedData[memberEmail]["notes"] = str(notes)
                                if rate != defaultRate:
                                    memberSavedData[memberEmail]["rate"] = float(rate)

                        initMembersFile(year)  # Réinit/Création de la liste des adhérents

                    newMember = Member(payment.email, payment.name, payment.surname, payment.address,
                                       payment.postalCode, payment.city, payment.phone)  # TODO : set by keys
                    if payment.email in memberSavedData:  # Si on a récup des remarques ou un tarif pour cet adh
                        if memberSavedData[payment.email]["notes"] is not None:
                            newMember.notes = memberSavedData[payment.email]["notes"]  # On enregistre les remarques pour l'instance
                        if memberSavedData[payment.email]["rate"] != defaultRate:
                            memberRate = memberSavedData[payment.email]["rate"]
                            if str(memberRate).replace('.', '').isnumeric():  # Si c'est une valeur numérique
                                newMember.rate = float(memberRate)  # On enregistre le tarif directement
                            else:  # Sinon on cherche si ça correspond à un nom de tarif enregistré
                                rate = Save().getRateByName(str(memberRate))
                                if rate is not None:
                                    newMember.rate = rate["value"]
                                else:
                                    LogManager().addLog("update", LogManager.LOGTYPE_WARNING,
                                                        f"Le tarif renseigné pour {newMember.name} {newMember.surname} est incorrect : '{memberRate}'.\nLe tarif par défaut sera utilisé à la place.")
                                    newMember.rate = Save().defaultRate["value"]

                    nbMember = getNbMemberInList(newMember.email, newMember.name, newMember.surname,
                                                 membersByYear[year])
                    if type(nbMember) is int:  # Si on a déjà l'adh pour l'année en question
                        member = membersByYear[year][nbMember]
                        member.updateContactData(payment.address, payment.postalCode, payment.city,
                                                 payment.phone)  # On met à jour ses coords de contact si besoin
                        member.addPayment(payment)  # Puis on ajoute son paiement à sa liste de paiements
                    else:
                        membersByYear[year].append(newMember)  # Sinon on ajoute le nouveau adh pour l'année en question
                        newMember.addPayment(payment)  # On ajoute le paiement à sa liste de paiements
            self.incrementProgress(labelTxt=f"{self.currentStep}/{self.nbSteps} : Traitement des fichiers de paiements")

        years.sort()  # On trie la liste par ordre chrn
        self.resetProgress()
        self.setNbSteps(len(years * 2) + 1)  # *2 Parce que d'abord on modifie les données par rapport aux années précédentes puis on exporte. +1 Pour le fichier de sauvegarde

        """
        Dans un second temps, on va regarder si pour chaque adhérent, celui-ci est enregistré dans l'une des années précédentes.
        Son statut ainsi que les valeurs des différents montants pourront être mis à jour.
        """

        for year in years:
            for member in membersByYear[year]:
                prevYear = int(year) - 1
                if len(membersByYear[prevYear]) > 0:
                    nbMember = getNbMemberInList(member.email, member.name, member.surname,
                                                 membersByYear[prevYear])
                    if type(nbMember) is int:
                        paidMembershipNextYear = membersByYear[prevYear][nbMember].amounts["paidMembershipNextYear"]
                        if paidMembershipNextYear > 0:
                            member.amounts["paidMembershipLastYear"] = paidMembershipNextYear
                            member.lastMembership = int(year) - 1

                total = member.amounts["totalYear"]
                restToPay = member.rate - member.amounts["paidMembershipLastYear"]
                membershipThisYear = min(restToPay, total)

                membershipNextYear = 0
                if len(member.receipts) > 0 and datetime.strptime(member.receipts[-1].date, "%d/%m/%Y").month >= 9 and total >= member.rate:
                    membershipNextYear = min(restToPay, total - membershipThisYear)
                    member.status = "RAA"

                totalDonation = total - membershipThisYear - membershipNextYear

                member.amounts["paidMembershipYear"] = membershipThisYear
                member.amounts["paidMembershipNextYear"] = membershipNextYear
                member.amounts["donationsYear"] = totalDonation

                if member.status == "NA" or member.status == "DON-ADH":
                    for yearPrec in years:
                        nbMember = getNbMemberInList(member.email, member.name, member.surname, membersByYear[yearPrec])
                        if type(nbMember) is int:
                            if yearPrec < year:
                                member.lastMembership = yearPrec
                                member.status = "RA"

            self.incrementProgress(labelTxt=f"{self.currentStep}/{self.nbSteps} : Mise à jour des adhérents")

        """Enfin, on exporte les listes des adhérents (xlsx) et les reçus (PDF) """
        for year in years:
            exportMembersFile(self.paths["listesAdherents"] / f"{year}.xlsx", membersByYear[year])
            exportMemberReceipts(membersByYear[year])
            self.incrementProgress(labelTxt=f"{self.currentStep}/{self.nbSteps} : Exportation de la liste des adhérents et des reçus")

        """ Puis on sauvegarde """
        Save().save()
        self.incrementProgress(labelTxt=f"{self.currentStep}/{self.nbSteps} : Enregistrement du fichier de sauvegarde")

        LogManager().addLog("update", LogManager.LOGTYPE_INFO, "Succès du traitement des fichiers de paiements.")
        self.resetProgress()  # TODO: stop

        return LogManager().getHigherStatusOf("update")

    def setNbSteps(self, nbSteps):
        self.nbSteps = nbSteps

    def incrementProgress(self, incrSteps=1, labelTxt=''):
        self.currentStep += incrSteps
        self.progressPercent = (self.currentStep-1)/self.nbSteps * 100
        if self.view:
            self.view.updateLbl.config(text='')
            self.view.progTxt.set(labelTxt)
            self.view.progVal.set(min(self.progressPercent, 99.9))  # Mise à jour de la variable liée à la barre de progression
            self.view.update_idletasks()  # Forcer la mise à jour de l'interface graphique

    def resetProgress(self):
        self.currentStep = 1
        self.progressPercent = 0
        if self.view:
            self.view.updateLbl.config(text='')
            self.view.progVal.set(0)  # Reset la barre de progression
            self.view.update_idletasks()  # Forcer la mise à jour de l'interface graphique
