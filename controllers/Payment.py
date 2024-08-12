import csv
import logging

from openpyxl import load_workbook

from models.Payment import Payment
from utils import utils


def getDataFromPaymentsFile(path, source):
    payments = []

    def addToPayments(payment):
        if newPayment.isValid:
            payments.append(payment)
        else:
            logging.error(f"Impossible d'ajouter le paiement '{payment.source.upper()}' venant de '{payment.name} {payment.surname}'. Ref: '{payment.refPayment}'.\n{payment.notValidCause}")

    if source != "cb":
        workbook = load_workbook(filename=path)
        sheet = workbook.active
    else:
        csvContent = []
        with open(path, mode='r') as csvFile:
            csvReader = csv.DictReader(csvFile, delimiter=',')
            for row in csvReader:
                csvContent.append(row)

    if source == "helloAsso":
        requiredCols = [1, 2, 5, 6, 7]  # Indices des colonnes à vérifier pour les valeurs non nulles
        for row in sheet.iter_rows(min_row=2):
            if all(row[idx].value is not None for idx in requiredCols):
                newPayment = Payment(
                        email       = row[7].value,
                        name        = row[5].value,
                        surname     = row[6].value,
                        date        = row[2].value.strftime("%d/%m/%Y"),
                        regular     = row[8].value == "Crowdfunding",
                        address     = row[9].value,
                        postalCode  = str(row[10].value),
                        city        = row[11].value,
                        phone       = '',
                        amount      = float(row[1].value),
                        source      = source,
                        refPayment  = row[0].value
                )
                addToPayments(newPayment)

    elif source == "paypal":
        requiredCols = [0, 2, 8, 9, 4]
        for row in sheet.iter_rows(min_row=4):
            if all(row[idx].value is not None for idx in requiredCols) and row[4].value == "Terminé" and float(row[8].value) > 0:
                names = row[2].value.rsplit(' ', 1)
                newPayment = Payment(
                        email       = row[9].value,
                        name        = names[0],
                        surname     = names[1],
                        date        = row[0].value.strftime("%d/%m/%Y"),
                        regular     = row[3].value == "Paiement d'abonnement",
                        address     = row[11].value,
                        postalCode  = str(row[13].value or ''),
                        city        = row[12].value,
                        phone       = '',
                        amount      = float(row[8].value),
                        source      = source,
                        refPayment  = row[10].value
                    )
                addToPayments(newPayment)

    elif source == "virEspChq":
        requiredCols = [0, 2, 3, 4, 8]
        for row in sheet.iter_rows(min_row=2):
            if all(row[idx].value is not None for idx in requiredCols):
                source = ''
                mode = row[9].value.casefold()
                if mode == 'v':
                    source = "Virement"
                elif mode == 'c':
                    source = "Chèque"
                elif mode == 'e':
                    source = "Espèce"

                newPayment = Payment(
                    email       = row[2].value,
                    name        = row[3].value,
                    surname     = row[4].value,
                    date        = row[0].value.strftime("%d/%m/%Y"),
                    regular     = row[10].value == "O",
                    address     = row[5].value,
                    postalCode  = str(row[6].value),
                    city        = row[7].value,
                    phone       = '',
                    amount      = float(row[8].value),
                    source      = source,
                    refPayment  = row[1].value
                )
                addToPayments(newPayment)

    elif source == "cb":
        requiredCols = ["Heure de soumission", "Nom", "Prénom", "E-mail", "Carte de crédit/débit - Montant",
                        "Carte de crédit/débit - État "]
        for row in csvContent:
            if all(row[key] is not None for key in requiredCols) and row["Carte de crédit/débit - État "].casefold() == "completed":
                newPayment = Payment(
                        email       = row["E-mail"],
                        name        = row["Nom"],
                        surname     = row["Prénom"],
                        date        = utils.convertFrenchDate(row["Heure de soumission"]).strftime("%d/%m/%Y"),
                        regular     = False,
                        address     = row["Address - Rue"] + ' ' + row["Address - Appartement, suite, etc."],
                        postalCode  = str(row["Address - Code postal"]),
                        city        = row["Address - Ville"],
                        phone       = str(row["Téléphone"]),
                        amount      = float(row["Carte de crédit/débit - Montant"]),
                        source      = source,
                        refPayment  = str(row["Carte de crédit/débit - ID de la transaction"])
                )
                addToPayments(newPayment)

    return payments