import csv
from glob import glob
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from pathlib import Path
import logging
from datetime import datetime

## Styles
widthCols = [16, 25, 15, 15, 15, 5, 20, 8, 25, 10, 25, 15, 8, 8, 8, 8, 8, 15, 10, 20]
fontHeader = Font(b=True, size=8)
sideHeader = Side(border_style="thin", color="000000")
borderHeader = Border(
    left=sideHeader, right=sideHeader, top=sideHeader, bottom=sideHeader
)

sideBody = Side(border_style="thin", color="808080")
borderBody = Border(
    left=sideBody, right=sideBody, top=sideBody, bottom=sideBody
)



## Utils
def convertFrenchDate(frenchDateStr):
    # Dictionnaire pour mapper les mois français aux mois anglais
    frenchToEnglishMonths = {
        "Janvier": "January", "Février": "February", "Mars": "March",
        "Avril": "April", "Mai": "May", "Juin": "June",
        "Juillet": "July", "Août": "August", "Septembre": "September",
        "Octobre": "October", "Novembre": "November", "Décembre": "December"
    }

    # Remplacer le mois français par le mois anglais
    for frenchMonth, englishMonth in frenchToEnglishMonths.items():
        if frenchMonth in frenchDateStr:
            englishDateStr = frenchDateStr.replace(frenchMonth, englishMonth)
            break

    # Convertir la chaîne modifiée en objet datetime
    return datetime.strptime(englishDateStr, "%B %d, %Y @ %I:%M %p")

## Fonctions

def createOrResetMembersList(year):
    membersList = paths["listesAdherents"]/f"{year}.xlsx"
    if not Path(membersList).is_file():
        try:
            workbook = Workbook()
            sheet = workbook.active

            fields = ["Enregistrement", "Adresse mail", "Nom", "Prenom", "IDs reçus", "Statut", "Date dernière adhésion", "Envoi mails", "Adresse", "Code postal", "Ville", "Téléphone", "Montant adhésion payé année n-1", "Montant adhésion année n", "Montant adhésion année n+1", "Montant dons année n", "Total année n", "Dernier paiement", "Dern. moy. paiement", "Remarques"]
            
            # Ajouter les en-têtes de colonnes
            sheet.append(fields)
            
            # repositionner le curseur
            sheet.cell(row=2, column=1)

            for nbCol, width in enumerate(widthCols):
                col = sheet.column_dimensions[get_column_letter(nbCol+1)]
                col.width = width
            
            # Appliquer le style de l'en-tête à la première ligne
            for cell in sheet[1]:
                cell.font = fontHeader
                cell.border = borderHeader
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Sauvegarder le fichier
            workbook.save(membersList)
            workbook.close()
            logging.info(f"Fichier {membersList} créé")
            return sheet
        except OSError as error:
            logging.error(f"Impossible de créer le fichier {membersList} : {error}")
            return False
    else: ## Sinon remettre tout à zero TODO: garder les remarques ? Mettre avec le fichier cache pour les reçus
        try:
            workbook = load_workbook(membersList)
            sheet = workbook.active

            sheet.delete_rows(2, sheet.max_row)
            workbook.save(membersList)
            workbook.close()
            return sheet
        except OSError as error:
                logging.error(f"Impossible d'ouvrir le fichier {membersList} : {error}")
                return False


def findMember(email, name, surname, membersArray):
    #D'abord chercher par mail
    for rowNb, row in enumerate(membersArray):
        if row[1] != None and row[1] == email:
            return rowNb+1
    #Puis chercher par prenom + nom
    for rowNb, row in enumerate(membersArray):
        if row[2] != None and row[2] != None and row[2].casefold() == name.casefold() and row[3].casefold() == surname.casefold():
            return rowNb+1
    return False


def getDataFromRecordsFile(path, source):
    if source != "cb":
        workbook = load_workbook(filename=path)
        sheet = workbook.active
    else:
        csvContent = []
        with open(path, mode='r') as csvFile:
            csvReader = csv.DictReader(csvFile, delimiter=',')
            for row in csvReader:
                csvContent.append(row)

    payments = [] 
    if source=="helloAsso":
        requiredCols = [1, 2, 5, 6, 7]  #Indices des colonnes à vérifier pour les valeurs non nulles
        for row in sheet.iter_rows(min_row=2):
            if all(row[idx].value is not None for idx in requiredCols):
                payments.append({
                    "mail": row[7].value,
                    "nom": row[5].value,
                    "prenom": row[6].value,
                    "date": row[2].value.strftime("%d/%m/%Y"), 
                    "type": "reguliers" if row[8].value == "Crowdfunding" else "ponctuels",
                    "adresse": row[9].value,
                    "cp": row[10].value,
                    "ville": row[11].value,
                    "telephone": None,
                    "montant": float(row[1].value),
                    "source": source
                })
    elif source=="paypal":
        requiredCols = [0, 2, 8, 9, 4]
        for row in sheet.iter_rows(min_row=4):
            if all(row[idx].value is not None for idx in requiredCols) and row[4].value == "Terminé" and float(row[8].value) > 0:
                prenomNom = row[2].value.split(' ', 1)
                payments.append({
                    "mail": row[9].value,
                    "nom": prenomNom[0],
                    "prenom": prenomNom[1],
                    "date": row[0].value.strftime("%d/%m/%Y"), 
                    "type": "reguliers" if row[3].value == "Paiement d'abonnement" else "ponctuels",
                    "adresse": row[11].value,
                    "cp": row[13].value,
                    "ville": row[12].value,
                    "telephone": None,
                    "montant": float(row[8].value),
                    "source": source
                })
    elif source=="virEspChq":
        requiredCols = [0, 2, 3, 4, 8]
        for row in sheet.iter_rows(min_row=2):
            if all(row[idx].value is not None for idx in requiredCols):
                source = ''
                mode = row[9].value.casefold()
                if mode == 'v':
                    source="Virement"
                elif mode == 'c':
                    source="Chèque"
                elif mode == 'e':
                    source="Espèce"
                payments.append({
                    "mail": row[2].value,
                    "nom": row[3].value,
                    "prenom": row[4].value,
                    "date": row[0].value.strftime("%d/%m/%Y"), 
                    "type": "reguliers" if row[10].value == "O" else "ponctuels",
                    "adresse": row[5].value,
                    "cp": row[6].value,
                    "ville": row[7].value,
                    "telephone": None,
                    "montant": float(row[8].value),
                    "source": source
                })
    elif source=="cb":
        requiredCols = ["Heure de soumission", "Nom", "Prénom", "E-mail", "Carte de crédit/débit - Montant", "Carte de crédit/débit - État "]
        for row in csvContent:
            if all(row[key] is not None for key in requiredCols) and row["Carte de crédit/débit - État "].casefold() == "completed":
                payments.append({
                    "mail": row["E-mail"],
                    "nom": row["Nom"],
                    "prenom": row["Prénom"],
                    "date": convertFrenchDate(row["Heure de soumission"]).strftime("%d/%m/%Y"), 
                    "type": "ponctuels",
                    "adresse": row["Address - Rue"] + ' ' + row["Address - Appartement, suite, etc."],
                    "cp": row["Address - Code postal"],
                    "ville": row["Address - Ville"],
                    "telephone": row["Téléphone"],
                    "montant": float(row["Carte de crédit/débit - Montant"]),
                    "source": source
                })

    return payments


#Configurer Logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

## Créer les dossiers si besoin ##
currentPath = Path.cwd()
dataPath = currentPath/"donnees"

currentDatetime = datetime.now()

paths = { #TODO : use Path()
    "actuel": dataPath,
    "listesAdherents": dataPath/"listesAdherents",
    "paiementsHelloAsso": dataPath/"paiements/helloAsso",
    "paiementsPaypal": dataPath/"paiements/paypal",
    "paiementsVirEspChq": dataPath/"paiements/virEspChq",
    "paiementsCB": dataPath/"paiements/CB",
    "recusFiscaux": dataPath/"recusFiscaux"/str(currentDatetime.year)/str(currentDatetime.month)
}

for key, path in paths.items():
    try:
        path.mkdir(parents=True, exist_ok=True)
        logging.info(f"Dossier créé ou déjà existant : {path}")
    except OSError as error:
        logging.error(f"Impossible de créer le dossier {path} : {error}")


paymentsFiles = {
    "helloAsso": glob(str(paths["paiementsHelloAsso"]/"*.xlsx")),
    "paypal": glob(str(paths["paiementsPaypal"]/"*.xlsx")),
    "virEspChq": glob(str(paths["paiementsVirEspChq"]/"*.xlsx")),
    "cb": glob(str(paths["paiementsCB"]/"*.csv"))
}


### Un sheet par année requise
recordsByYear = {}

records = []
for source, filePaths in paymentsFiles.items():
    for filePath in filePaths:
        records.append(getDataFromRecordsFile(filePath, source))

years = []

for source in records:
    for record in source:
        record["year"] = str(datetime.strptime(record["date"], "%d/%m/%Y").year)
        if not record["year"] in years: #Si c'est la première fois à l'execution qu'on tombe sur cette année
            years.append(record["year"])
            recordsByYear[record["year"]] = []
            sheet = createOrResetMembersList(record["year"])
          #  nbRow = False #Pas possible de maj un membre comme il n'y en a pas
        #else:
         #   nbRow = findMember(record["mail"], record["nom"], record["prenom"], recordsByYear[record["year"]]) #On peut chercher si le membre existe déjà

        nbRow = findMember(record["mail"], record["nom"], record["prenom"], recordsByYear[record["year"]]) #On peut chercher si le membre existe déjà
        currMembersList = recordsByYear[record["year"]]
        if nbRow:
            #maj
            #Vérifier si adresse identique
            if currMembersList[nbRow][8] == None or record["adresse"] != None and currMembersList[nbRow][8].casefold() != record["adresse"].casefold(): #Nouvelle adresse
                currMembersList[nbRow][8] = record["adresse"]
            if currMembersList[nbRow][9] == None or record["cp"] != None and str(currMembersList[nbRow][9]).casefold() != str(record["cp"]).casefold(): #Nouveau CP 
                currMembersList[nbRow][9] = str(record["cp"])
            if currMembersList[nbRow][10] == None or record["ville"] != None and currMembersList[nbRow][10].casefold() != record["ville"].casefold(): #Nouvelle ville
                currMembersList[nbRow][10] = record["ville"]
            #Vérifier si téléphone identique
            if currMembersList[nbRow][11] == None or record["telephone"] != None and currMembersList[nbRow][10] != record["telephone"]:
                currMembersList[nbRow][11] = record["telephone"]

            #Mettre à jour la date du dernier paiement + sa source si paiement plus récent que ce qui a été enregistré
            if currMembersList[nbRow][17] == None or datetime.strptime(currMembersList[nbRow][17], "%d/%m/%Y") < datetime.strptime(record["date"], "%d/%m/%Y"):
                currMembersList[nbRow][17] = record["date"]
                currMembersList[nbRow][18] = record["source"]
            #Ajouter le montant
            newAmount = int(currMembersList[nbRow][15]) + record["montant"] #rename
            membershipAmount = 18
            currMembersList[nbRow][13] = min(newAmount, membershipAmount) #Montant adh n
            newAmount -= currMembersList[nbRow][13]
            status = "DON-ADH"
            if datetime.strptime(record["date"], "%d/%m/%Y").month >= 9: #Après septembre
                currMembersList[nbRow][14] = min(newAmount, membershipAmount) #Montant adh n+1
                newAmount -= currMembersList[nbRow][14]
                status = "RAA"
            currMembersList[nbRow][5] = status
            currMembersList[nbRow][15] = max(0, newAmount) #Montant dons
            currMembersList[nbRow][16] = currMembersList[nbRow][15] + currMembersList[nbRow][13] + currMembersList[nbRow][14] #total (adhs + dons)
        else:
            #create
            newAmount = record["montant"]
            amountMembershipY = min(newAmount, 18) #montant adh n
            newAmount -= amountMembershipY
            amountMembershipY1 = 0
            status = "NA"
            if datetime.strptime(record["date"], "%d/%m/%Y").month >= 9: #Après septembre
                amountMembershipY1 = min(newAmount, 18) #montant adh n+1
                newAmount -= amountMembershipY1
                status = "RAA"
            donationsAmount = max(0, newAmount) #montant dons

            fields = [currentDatetime.strftime("%d/%m/%Y"), 
                            record["mail"], record["nom"], 
                            record["prenom"], None, status, None, "Oui", 
                            record["adresse"], record["cp"], 
                            record["ville"], record["telephone"], 
                            0,
                            amountMembershipY, amountMembershipY1,
                            donationsAmount, record["montant"], 
                            record["date"], record["source"], '/']

            currMembersList.append(fields)


years.sort(reverse=True)

# Appliquer le style du corps aux autres lignes
for year in recordsByYear:
    recordsByYear[year] = sorted(recordsByYear[year], key=lambda cols: datetime.strptime(cols[17], "%d/%m/%Y"))
    workbookPath = paths["listesAdherents"]/f"{year}.xlsx"
    workbook = load_workbook(filename=workbookPath)
    sheet = workbook.active
    for record in recordsByYear[year]: #adh, pas record
        # Vérifier si le donateur a déjà adhéré
    
        for y in years:
            nbMember = findMember(record[1], record[2], record[3], recordsByYear[y])
            if int(y) < int(year) and nbMember:
                record[6] = y
                if record[5] == "NA":
                    record[5] = "RA"
                #if int(y) == int(year)-1:
                    
                #break

        sheet.append(record)
        
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = borderBody
    
    workbook.save(workbookPath)
    workbook.close()
