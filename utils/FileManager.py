import csv
import io
import os
import re
from datetime import datetime
from operator import attrgetter
from os import path
from pathlib import Path
from shutil import rmtree

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter
from pypdf import generic

from models.Payment import Payment
from models.Save import Save
from utils import styles, misc
from utils.LogManager import LogManager
from utils.PathManager import PathManager
from models.Thunderbird import Thunderbird


def getDataFromPaymentsFile(path, source):  # todo : A découper en plusieurs morceaux
    payments = []

    def isEmptyRow(row):
        """
        Vérifie si une ligne est complètement vide.
        """
        return all(cell.value is None or cell.value == '' for cell in row)

    def getMissingRequiredCol(row, requiredCols):
        missingCols = []
        for colName, colIdx in requiredCols.items():
            if row[colIdx].value is None:
                missingCols.append(colName)
        return missingCols

    def addWarningMissingCols(nbRow, missingCols):
        LogManager().addLog("update", LogManager.LOGTYPE_WARNING,
                            f"Fichier {path}.\nIl manque des valeurs obligatoires pour la ligne n°{nbRow} :\nValeur(s) manquante(s) : {', '.join(missingCols)}")

    def addToPayments(payment):
        if newPayment.isValid:
            payments.append(payment)
        else:
            LogManager().addLog("update", LogManager.LOGTYPE_WARNING,
                                f"Impossible de valider le paiement '{payment.source.upper()}' venant de {payment.lastName} {payment.firstName}. Référence du paiement : '{payment.refPayment}'.\nRaison -> {payment.notValidCause}")
    csvArray = []
    if source != "cb":
        workbook = load_workbook(filename=path)
        sheet = workbook.active
    else:
        header = ["Heure de soumission", "Nom", "Prénom", "E-mail", "Téléphone", "Address - Rue",
                  "Address - Appartement suite etc.", "Address - Ville", "Address - Code postal", "Nationalité",
                  "Profession", "Date de naissance", "Est-ce une ré-adhésion ?", "Montant don",
                  "Carte de crédit/débit - Mode", "Carte de crédit/débit - Nom du produit / forfait",
                  "Carte de crédit/débit - Type de paiement", "Carte de crédit/débit - Montant",
                  "Carte de crédit/débit - Devise", "Carte de crédit/débit - Quantité",
                  "Carte de crédit/débit - ID de la transaction", "Carte de crédit/débit - État ",
                  "Carte de crédit/débit - Gérer"]

        with open(path, mode='r', newline='', encoding="utf-8-sig") as file:
            csvContent = file.read()
            lines = []
            for line in csvContent.splitlines():
                # 0) Replace Appartement, suite, etc
                line = line.replace("Appartement, suite, etc", " Appartement suite etc")

                # 1) Remove " from beginning and end of each line
                if line.startswith('"') and line.endswith('"'):
                    line = line[1:-1]
                # 2) Remove the , from the datetime
                line = re.sub(
                    r'(\s[0-9]{1,2}),(?=\s[0-9]{4}\s@\s[0-9]{1,2}:[0-9]{2}\s(?:PM|AM))',
                    r'\1',
                    line
                )

                # 3) Remove , inside ""
                line = re.sub(r'""([^"]+)""', lambda m: '""' + m.group(1).replace(',', '') + '""', line)

                # 4) Replace "" by nothing
                line = line.replace('""', '')

                if line.count(',') == 21: # Missing a col
                    line = re.sub(
                        r'(,\s*[0-9\s]{10,}\s*,[^,]+)',
                        r'\1,',
                        line
                    )

                lines.append(line)

                reader = csv.DictReader(io.StringIO('\n'.join(lines)), delimiter=',')
                rows = list(reader)

    try: # todo: A optimiser
        if source == "helloAsso":
            requiredCols = {"montant":1, "date":2, "nom de famille":5, "prénom":6, "adresse mail":7}
            for row in sheet.iter_rows(min_row=2):
                if isEmptyRow(row):
                    continue  # Si ligne vide, on ignore et on passe à la suivante
                missingCols = getMissingRequiredCol(row, requiredCols)
                if missingCols:
                    addWarningMissingCols(row[0].row, missingCols)
                    continue  # Si ligne manque des valeurs obligatoires, on affiche un message et on ignore la ligne

                newPayment = Payment(
                    email=row[7].value,
                    lastName=row[5].value,
                    firstName=row[6].value,
                    date=row[2].value.strftime("%d/%m/%Y"),
                    regular=row[9].value == "Don mensuel",
                    address=row[10].value,
                    postalCode=str(row[11].value),
                    city=row[12].value,
                    phone='',
                    amount=float(row[1].value),
                    source=source,
                    refPayment=row[0].value
                )
                addToPayments(newPayment)

        elif source == "paypal":
            requiredCols = {"date":0, "noms":2, "montant":8, "adresse mail":9, "état":4}
            for row in sheet.iter_rows(min_row=2):
                if row[4].value != "Terminé" or float(row[8].value) == 0 or row[3].value not in ("Paiement de don", "Paiement d'abonnement"):
                    continue
                if isEmptyRow(row):
                    continue  # Si ligne vide, on ignore et on passe à la suivante
                missingCols = getMissingRequiredCol(row, requiredCols)
                if missingCols:
                    addWarningMissingCols(row[0].row, missingCols)
                    continue  # Si ligne manque des valeurs obligatoires, on affiche un message et on ignore la ligne

                names = row[2].value.rsplit(' ', 1)
                if len(names) != 2:
                    LogManager().addLog("update", LogManager.LOGTYPE_WARNING,
                                        f"Pour le paiement n°{row[10].value} : Il y a une erreur avec le nom du donateur '{''.join(names)}'")
                    continue

                newPayment = Payment(
                    email=row[9].value,
                    lastName=names[1],
                    firstName=names[0],
                    date=row[0].value.strftime("%d/%m/%Y"),
                    regular=row[3].value == "Paiement d'abonnement",
                    address=row[11].value,
                    postalCode=str(row[13].value or ''),
                    city=row[12].value,
                    phone='',
                    amount=float(row[6].value),
                    source=source,
                    refPayment=row[10].value
                )
                addToPayments(newPayment)

        elif source == "virEspChq":
            requiredCols = {"date":0, "adresse mail":2, "nom de famille":3, "prénom":4, "montant":8, "source":9}
            for row in sheet.iter_rows(min_row=2):
                if isEmptyRow(row):
                    continue  # Si ligne vide, on ignore et on passe à la suivante
                missingCols = getMissingRequiredCol(row, requiredCols)
                if missingCols:
                    addWarningMissingCols(row[0].row, missingCols)
                    continue  # Si ligne manque des valeurs obligatoires, on affiche un message et on ignore la ligne

                source = ''
                mode = row[9].value.casefold()
                if mode == 'v':
                    source = "Virement"
                elif mode == 'c':
                    source = "Chèque"
                elif mode == 'e':
                    source = "Espèce"

                newPayment = Payment(
                    email=row[2].value,
                    lastName=row[3].value,
                    firstName=row[4].value,
                    date=row[0].value.strftime("%d/%m/%Y"),
                    regular=row[10].value == "O",
                    address=row[5].value,
                    postalCode=str(row[6].value),
                    city=row[7].value,
                    phone='',
                    amount=float(row[8].value),
                    source=source,
                    refPayment=row[1].value
                )
                addToPayments(newPayment)

        elif source == "cb":
            for row in rows:
                newPayment = Payment(
                    email=row["E-mail"],
                    lastName=row["Nom"],
                    firstName=row["Prénom"],
                    date=misc.convertFrenchDate(row["Heure de soumission"]).strftime("%d/%m/%Y"),
                    regular=False,
                    address=row["Address - Rue"] + ' ' + row["Address -  Appartement suite etc."],
                    postalCode=str(row["Address - Code postal"]),
                    city=row["Address - Ville"],
                    phone=str(row["Téléphone"]),
                    amount=float(row["Carte de crédit/débit - Montant"]),
                    source=source,
                    refPayment=str(row["Carte de crédit/débit - ID de la transaction"])
                )
                addToPayments(newPayment)

    except Exception as e:
        LogManager().addLog("update", LogManager.LOGTYPE_ERROR, f"Une erreur est survenue lors du traitement du fichier {source} : {path}. Veuillez vérifier la structure du fichier et le format des données. (erreur rencontrée : {e})")

    return payments

def initMembersFile(year):  # Création du fichier ou réinitialisation
    membersList = PathManager().getPaths()["listesAdherents"] / f"liste des adhérents {year}.xlsx"
    if not Path(membersList).is_file():
        try:
            year = str(year)
            precYear = str(int(year) - 1)
            nextYear = str(int(year) + 1)

            workbook = Workbook()
            sheet = workbook.active

            fields = ["Adresse mail", "Nom", "Prénom", "IDs reçus", "Statut",
                      "Dern. adh.", "Régulier", "Envoi mails", "Adresse", "Code postal", "Ville", "Téléphone",
                      "Adh. " + precYear, "Adh. " + year, "Adh. " + nextYear,
                      "Dons " + year, "Total " + year, "Dern. paiement", "Dern. moy. paiement", "Tarif", "Remarques"]

            # Ajouter les en-têtes de colonnes
            sheet.append(fields)

            # repositionner le curseur
            sheet.cell(row=2, column=1)

            for nbCol, width in enumerate(styles.widthCols):
                col = sheet.column_dimensions[get_column_letter(nbCol + 1)]
                col.width = width

            # Appliquer le style de l'en-tête à la première ligne
            for cell in sheet[1]:
                cell.font = styles.fontHeader
                cell.border = styles.borderHeader
                cell.alignment = styles.centerAlign

            # Sauvegarder le fichier
            workbook.save(membersList)
            workbook.close()
            LogManager().addLog("update", LogManager.LOGTYPE_INFO, f"Fichier {membersList} créé.")
            return sheet
        except OSError as error:
            LogManager().addLog("error", LogManager.LOGTYPE_ERROR,
                                f"Impossible de créer le fichier {membersList} : {error}.")
            return False
    else:  # Sinon remettre tout à zero
        try:
            workbook = load_workbook(membersList)
            sheet = workbook.active

            sheet.delete_rows(2, sheet.max_row)
            workbook.save(membersList)
            workbook.close()
            return sheet
        except OSError as error:
            LogManager().addLog("error", LogManager.LOGTYPE_ERROR,
                                f"Impossible de créer le fichier {membersList} : {error}.")
            return False


def getExistingMembersData(year):  # Permet de récupérer les éventuels notes et tarifs spéciaux dans une liste d'adhérents (donnée par une année)
    filepath = PathManager().getPaths()["listesAdherents"] / f"liste des adhérents {year}.xlsx"
    existingMembersData = {}
    if Path(filepath).is_file():  # Si le fichier existe déjà
        existingMembersData = {}  # Pour récupérer les éventuels remarques et tarifs spéciaux
        defaultRate = Save().defaultRate["value"]

        workbook = load_workbook(filepath)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_row=len(sheet['A']) - 4):  # On regarde s'il y a des remarques à récupérer pour les remettre dans le fichier final si l'adhérent y est toujours présent
            memberEmail = row[0].value
            notes = row[20].value
            rate = row[19].value
            existingMembersData[memberEmail] = {}
            if notes is not None:
                existingMembersData[memberEmail]["notes"] = notes
            if rate != defaultRate:
                existingMembersData[memberEmail]["rate"] = rate

    return existingMembersData

def exportMembersFile(filePath, members):
    def calcTotalsPayments(members):
        totals = {
            "withReceipt": 0,
            "withoutReceipt": 0,
            "nbReceipts": 0
        }

        for member in members:
            # Handle regular payment receipt
            if member.regularPaymentsReceipt and not member.regularPaymentsReceipt.canBeExported:
                totals["withoutReceipt"] += member.regularPaymentsReceipt.amount

            # Process individual receipts
            totals["withReceipt"] += sum(receipt.amount for receipt in member.receipts if receipt.canBeExported)
            totals["nbReceipts"] += sum(1 for receipt in member.receipts if receipt.canBeExported)

            # Add non-exported amounts
            totals["withoutReceipt"] += member.amountNotExported

        return totals

    try:
        workbook = load_workbook(filename=filePath)
        sheet = workbook.active

        members = sorted(list(members.values()), key=attrgetter("lastPayment"))  # Trie les membres selon la date du dernier paiement, par ordre chronologique
        for member in members:  # On ajoute chaque membre au fichier Excel
            sheet.append(member.toArray())

        # Calculer et afficher les totaux pour les paiements
        totals = calcTotalsPayments(members)
        startTotalsRow = len(sheet['A']) + 2
        sheet.cell(startTotalsRow, 1).value = "Nombre de reçus"
        sheet.cell(startTotalsRow, 2).value = totals["nbReceipts"]
        sheet.cell(startTotalsRow + 1, 1).value = "Total avec reçus"
        sheet.cell(startTotalsRow + 1, 2).value = totals["withReceipt"]
        sheet.cell(startTotalsRow + 2, 1).value = "Total sans reçus"
        sheet.cell(startTotalsRow + 2, 2).value = totals["withoutReceipt"]

        for row in sheet.iter_rows(min_row=2):  # Et pour le body du tableau, on applique le style correspondant
            for cell in row:
                cell.border = styles.borderBody

        workbook.save(filePath)  # Puis on sauvegarde
        workbook.close()
        LogManager().addLog("update", LogManager.LOGTYPE_INFO, f"Succès de l'exportation du fichier {filePath}")
    except(Exception,) as error:
        LogManager().addLog("update", LogManager.LOGTYPE_ERROR,
                            f"Une erreur est survenue lors de l'exportation du fichier {filePath}.\n{error}")


def exportMemberReceipts(members, progressBar):
    for email, member in members.items():
        if member.hasValidAddress():
            try:  # Exportation des PDFs
                receipts = member.receipts
                if member.regularPaymentsReceipt is not None:
                    receipts.append(member.regularPaymentsReceipt)
                exportedReceipts = _exportReceipts(receipts)

                try:  # Sauvegarde des reçus exportés avec succès
                    Save().addMemberReceipt(member.email, exportedReceipts)
                    progressBar.incrementProgress(
                        labelTxt=f"Exportation des reçus fiscaux pour chaque adhérent", showStep=True,
                        hideAfterFinish=False)

                except(Exception,) as error:
                    LogManager().addLog("update", LogManager.LOGTYPE_ERROR,
                                        f"Une erreur est survenue lors de l'enregistrement de {member.lastName} {member.firstName} dans le fichier de sauvegarde : {error}")

            except(Exception,) as error:
                LogManager().addLog("update", LogManager.LOGTYPE_ERROR,
                                    f"Une inconnue est survenue lors de l'exportation des reçus de {member.lastName} {member.firstName} : {error}'")

        else:
            LogManager().addLog("update", LogManager.LOGTYPE_WARNING,
                                f"{member.firstName} {member.lastName}  n'a pas de coordonnées de contact valides. L'édition de ses reçus est impossible.")


def _exportReceipts(receipts):  # TODO : Vérifier si erreur avant de faire exportedReceipts.append(receipt)
    paths = PathManager().getPaths()

    reader = PdfReader(paths["assets"]["PDFTemplate"])
    writer = PdfWriter(clone_from=reader)
    # writer.set_need_appearances_writer(True)

    writerItems = writer.get_fields().items()

    for k, v in writerItems:  # readonly
        o = v.indirect_reference.get_object()
        if o["/FT"] != "/Sig":
            o[generic.NameObject("/Ff")] = generic.NumberObject(o.get("/Ff", 0) | 1)

    exportedReceipts = []
    for receipt in receipts:
        if receipt.canBeExported:
            receiptDate = datetime.strptime(receipt.date, "%d/%m/%Y")
            receiptYear, receiptMonth = str(receiptDate.year), str(receiptDate.month)
            currentMonth = datetime.now().month
            if receipt.regular and currentMonth==12:
                directory = path.join(paths["recusFiscaux"], receiptYear, "reguliers")
            else:
                directory = path.join(paths["recusFiscaux"], receiptYear, receiptMonth)

            isDirOK = True
            if not path.isdir(directory):
                try:
                    Path(directory).mkdir(parents=True, exist_ok=True)
                    LogManager().addLog("update", LogManager.LOGTYPE_INFO, f"Dossier créé : {directory}")
                except OSError as error:
                    LogManager().addLog("update", LogManager.LOGTYPE_ERROR,
                                        f"Impossible de créer le dossier {directory} : {error}")
                    isDirOK = False

            if isDirOK:
                filepath = path.join(directory, f"{receipt.id}.pdf")
                # Vérifier si le reçu existe déjà et qu'il est identique. Doit également vérifier s'il existe bien dans le répertoire.
                receiptData = receipt.getDataDict()
                newHash = receipt.getHash()
                savedHash = None
                if path.isfile(filepath):
                    savedHash = Save().getSavedReceiptHash(receipt.member.email, receipt.id)
                if savedHash != newHash:
                    writer.update_page_form_field_values(
                        writer.pages[0],
                        receiptData,
                        auto_regenerate=False
                    )

                    with open(filepath, "wb") as output_stream:
                        writer.write(output_stream)
                        LogManager().addLog("update", LogManager.LOGTYPE_INFO, f"Création du reçu '{receipt.id}'.")

                exportedReceipts.append(receipt)
    return exportedReceipts


def importMembers():  # S'il y a plusieurs fois le même membre dans plusieurs listes, on fusionne ses ID de reçus
    PathManager().update()
    paths = PathManager().getPaths()
    memberListsPattern = paths["memberListsPattern"]

    members = {}

    for listPath in memberListsPattern:
        workbook = load_workbook(filename=listPath)
        sheet = workbook.active
        requiredCols = [0, 1, 2, 3]  # Indices des colonnes à vérifier pour les valeurs non nulles
        year = int(str(sheet.cell(row=1, column=14).value).replace("Adh. ", ''))
        for row in sheet.iter_rows(min_row=2):
            if all(row[idx].value is not None for idx in requiredCols):
                email = row[0].value
                IDReceipts = row[3].value.split(';')
                if email not in members:
                    member = {
                        "lastName": row[1].value,
                        "firstName": row[2].value,
                        "IDReceipts": IDReceipts,
                        "years": [year]
                    }
                    members[email] = member
                else:
                    members[email]["IDReceipts"] = list(set(members[email]["IDReceipts"] + IDReceipts))  # todo append ?
                    members[email]["years"].append(year)

        workbook.close()

    return members


def importReceipts():  # Obtient les données venant de Save + Thunderbird (pour les status des mails)
    thunderbirdEmails = Thunderbird().getStatusEmails()
    allMembersInSave = Save().members
    receipts = {}
    for email, member in allMembersInSave.items():
        receipts[email] = {}
        if "receipts" in member:
            for id, receipt in member["receipts"].items():
                datetimeReceipt = datetime.strptime(id[:6], "%y%m%d")
                dateStrReceipt = datetimeReceipt.strftime("%d/%m/%Y")
                receipt["date"] = dateStrReceipt
                receipts[email][id] = receipt
                if id in thunderbirdEmails:
                    statusID = thunderbirdEmails[id]
                else:
                    statusID = 0

                if statusID == 0:
                    statusTxt = "Non préparé"
                elif statusID == 1:
                    statusTxt = "Préparé"
                elif statusID == 2:
                    statusTxt = "Envoyé"
                elif statusID == 3:
                    statusTxt = "Supprimé"
                else:
                    statusTxt = '?'

                receipts[email][id]["emailStatus"] = statusTxt

    return receipts

def deleteAllReceipts():
    path = PathManager().getPaths()["recusFiscaux"]
    try:
        for directory in next(os.walk(path))[1]:
            rmtree(path / directory)
        return True
    except Exception as e:
        LogManager().addLog("OS", LogManager.LOGTYPE_ERROR, f"Impossible de supprimer le dossier : {e}")
        return False

def deleteAllMemberLists():
    path = PathManager().getPaths()["listesAdherents"]
    try:
        for file in next(os.walk(path))[2]:
            os.unlink(path / file)
        return True
    except Exception as e:
        LogManager().addLog("OS", LogManager.LOGTYPE_ERROR, f"Impossible de supprimer le dossier : {e}")
        return False
